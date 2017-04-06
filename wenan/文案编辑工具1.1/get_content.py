from openpyxl import Workbook
from openpyxl import load_workbook
import time
import random
import string
import math

main_list = []
dom_list = []
oth_list = []
content_list = []

wb = load_workbook(filename='wenan.xlsx')
ws = wb.get_sheet_by_name('Sheet1')


rows = ws.max_row
cols = ws.max_column
count = 0

def get_column(index,list):
    for i in range(2,rows+1):
            value = ws.cell(row=i,column=index).value
            if value==None:
                    break
            list.append(value)
	
	
def print_list(list):
    for item in list:
            print(item)
	
	
def rand_str():
    len = random.randint(6,8)	
    res = ''.join(random.sample(string.ascii_letters + string.digits, len))
    return res
		
get_column(1,main_list)
get_column(2,dom_list)
get_column(3,oth_list)

main_len = len(main_list)
dom_len = len(dom_list)
oth_len = len(oth_list)



def get_url(dom):
    str = '%s.baidu/www.%s'%(rand_str(),dom)
    return str

def generate():
    total_rec = math.ceil(dom_len*oth_len/2)
    for i in range(0,total_rec): 
        content = ""
        main_content = main_list[i%main_len]
        dom_content = get_url(dom_list[i%dom_len])
        oth_content = oth_list[i%oth_len]
        content += main_content
        content += '\n'
        content += dom_content
        content += '\n'
        content += '---------------------------------------'
        content += '\n'
        content += oth_content
        content_list.append(content)
    return total_rec



def genate_text():
    global count
    content += main_list[count%len(main_list)]
    content += '\n'
    dom = dom_list[count%len(dom_list)]
    url = get_url(dom)
    content += url
    content += '\n'
    content += '---------------------------------------\n'
    content += oth_list[count%len(oth_list)]
    count += 1
    content_list.append(content)
    return content


def save(list):
    wb = Workbook()
    ws1 = wb.active
    for val in list:
        rows = list.index(val)+1
        ws1.cell(row=rows,column=1).value = val

    time_stamp = time.strftime('%Y%m%d%H%M%S',time.localtime(time.time()))
    dest_file = '%s.xlsx'%(time_stamp)
    wb.save(filename=dest_file)



print('begin')

rec = generate()

print("Total record:%d"%(rec))

save(content_list)
print('done')
time.sleep(1)

        




