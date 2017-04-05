#coding=utf-8

import os
import re
import requests
import codecs
from bs4 import BeautifulSoup
import time
import urllib.parse
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import random
import string


headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36'}

#法律
url1 = 'https://zhidao.baidu.com/list?cid=103101&tag=%B7%A8%C2%C9'
#编程语言
url2 = 'https://zhidao.baidu.com/list?cid=110106&tag=%B1%E0%B3%CC%D3%EF%D1%D4'
#电脑网络
url3='https://zhidao.baidu.com/list?cid=110'
#操作系统
url4 = 'https://zhidao.baidu.com/list?cid=110103&tag=%B2%D9%D7%F7%CF%B5%CD%B3'
#感情
url6 = 'https://zhidao.baidu.com/list?cid=114102'


url_list = [url1,url2,url3,url4]
url_count = 0



eng_list = []
main_list = []
dom_list = []
oth_list = []
content_list = []
ans_count = 0


wb_ans = load_workbook(filename='wenan.xlsx')
ws_ans = wb_ans.get_sheet_by_name('Sheet1')
rows = ws_ans.max_row
cols = ws_ans.max_column

def get_column(index,list):
	for i in range(2,rows+1):
		value = ws_ans.cell(row=i,column=index).value
		if value==None:
			break
		list.append(value)
		
		
def rand_str():
	len = random.randint(6,8)	
	res = ''.join(random.sample(string.ascii_letters + string.digits, len))
	return res	


get_column(1,eng_list)
get_column(2,main_list)
get_column(3,dom_list)
get_column(4,oth_list)	



def genate_anwser():
	global ans_count
	content = ""
#	content += eng_list.pop()
#	content += '\n'
	content += main_list[ans_count%len(main_list)]
	content += '\n'
	content += rand_str()
	content += '.baidu/'
	content += '\n'
	content += '---------------------------------------\n'
	content += oth_list[ans_count%len(oth_list)]
	ans_count += 1
#	content_list.append(content)
	return content
	

def rest(i,j):
	t = random.randint(i,j)
	time.sleep(t)


def hide_my_anwser():
	browser.get(home)	
	#设置元素为可见,可以接受三种类型的参数：id(#)，class(.)，标签，很像jquery的选择器。
	js = 'document.querySelectorAll(".hide-my-reply")[0].style.display="block";'
	browser.execute_script(js)
	browser.find_element_by_class_name("hide-my-reply").click()
	rest(1,3)


def read_para(key):
	with open('config.txt','r') as f:
		for line in f.readlines():
			str_list = line.split('=')
			if key==str_list[0]:
				return str_list[1].strip('\n')
	f.close()
	print("cann't found %s"%(key))
	return ""


#检查回答数是多少
def check_question(url):
	doc = requests.get(url,headers=headers).content
	soup = BeautifulSoup(doc,'html.parser')
	tag_list = soup.find_all('div',class_="hd line other-hd")
	is_solved = soup.find('span',text=re.compile('最佳答案'))
	return len(tag_list) and is_solved

	
	
#如果找到完善回答表示回答成功，返回0，否则返回1
def check_result():
	doc =browser.page_source
	soup = BeautifulSoup(doc,'html.parser')
	res = soup.find_all('a',attrs={"alog-action":"qb-modify-ans"})
	if res:
		return 0
	else:
		return -1

#检查登录
def check_login():
	browser.get(home)
	try:
		browser.find_element_by_id('userbar-login')
		print("未登录或登录失败，请检查用户名和密码")
		exit(-1)
	except Exception :
		print(browser.get_cookies())
		print("login successful!")
		return 0

##########################################


#通过cookies登录
def log_by_cookie(cook):
	browser.get(home)
	cookie = browser.get_cookies()
	browser.add_cookie({'name':'BDUSS','value':cook})
	check_login()
	

#从文本文件中读取数据到list
def read_file(list,file):
	with open(file,'r') as f:
		for line in f.readlines():
			line = line.strip('\n')
			list.append(line)
	f.close()
	return list

	
	
#获取养号链接 
def add_question(list):
	global question_url
	count = 0
#	rand = random.randint(0,len(url_list)-1)
#	url = url_list[rand]
	url = url6
#	doc = requests.get(url,headers=headers).content
	browser.get(url)
	doc =browser.page_source
	soup = BeautifulSoup(doc,'html.parser')
	tag_list = soup.find_all('a',href=re.compile('https://zhidao.baidu.com/question/[\d]+.html'))

	for tag in tag_list:
		str = re.match('(?P<html_p>https://zhidao.baidu.com/question/[\d]+.html)',tag['href'])
		html = str.group('html_p')
		count += 1
		if(count==10):
			break
		list.append(html)
	print('抓取了%d个问题'%(len(list)))

	return ""


#获取提问标题	
def get_title(url):
	doc = requests.get(url,headers=headers).content

	soup = BeautifulSoup(doc,'html.parser')
	tag = soup.find(class_="ask-title ")#注意是class_
	if tag:
		title = tag.text.encode('gbk','ignore').decode('gbk','ignore')
		return title
	else:
		return None
	

#获取答案,如果没任何结果则返回None
def get_answer(title):
	prefix = u"https://zhidao.baidu.com/search?word="
	url = prefix + urllib.parse.quote(title)
	doc = requests.get(url,headers=headers).content

	soup = BeautifulSoup(doc,'html.parser')
	tag_list = soup.find_all('dd',class_="dd answer")
	for tag in tag_list:
		str = tag.text
		if len(str)>30:
			str = str[0:60]
			return str
	return None
	
	
	
#将list里面的链接保持到表格中
def save(list):
	wb = Workbook()
	ws1 = wb.active
	for val in list:
		rows = list.index(val)+1
		title = get_title(val)
		ws1.cell(row=rows,column=1).value = val
		#get_title可能需要处理一下
		ws1.cell(row=rows,column=2).value = title

	time_stamp = time.strftime('%Y%m%d%H%M%S',time.localtime(time.time()))
	dest_file = 'yanghao%s.xlsx'%(time_stamp)
	wb.save(filename=dest_file)

	
	
##########################################通过用户名和密码登录
def __login(username,password):
	log_url = 'https://passport.baidu.com/v2/?login&u='
	browser.get(log_url)
	user = browser.find_element_by_name("userName")
	user.clear()
	user.send_keys(username)
	passwd = browser.find_element_by_name("password")
	passwd.clear()
	passwd.send_keys(password)
	print('please login manually then press enter! thanks!')
	input()
#	browser.find_element_by_id("TANGRAM__PSP_3__submit").click()
#	time.sleep(5)
	return check_login()


	

	
##########################################检查验证码
def __authcode():
	global authcode
	try:
		browser.find_element_by_class_name("ik-authcode-input")
		print('please enter authcode:')
		code = input()
		rest(1,3)
		item = browser.find_element_by_class_name("ik-authcode-input")
		item.clear()
		item.send_keys(code)
		rest(1,3)
		authcode = 1
	except Exception:
		pass



#返回值 1表示出现异常，0表示回答成功，-1表示回答失败
def anwser_question(question_url):
	res = check_question(question_url)
	if res:
		print('跳过问题')
		return 1
	global coutinue_count,failed_status,status_count,count
	
	title = get_title(question_url)
	if not title:
		print('链接打不开')
		return 1
	print('养号：%s'%(title))
	
	browser.get(question_url)
	rest(1,3)
	#如果已经回答过，返回-1
	try:
		browser.find_element_by_xpath('//a[@alog-action="qb-modify-ans"]')
		print('回答过了')
		return 1
	except Exception:
		pass
	
		#如果没有提交回答，则先点击我有更好的答案
	try:
		browser.find_element_by_xpath('//a[@alog-action="qb-ans-sb"]') 
	except Exception:
		try:
			browser.find_element_by_id("answer-bar").click()
		except Exception:
				print('无法回答')
				return 1	

	
	answer = get_answer(title)
	if answer==None:
		print('找不到答案，跳过')
		return 1
	answer = answer
	browser.switch_to_frame('ueditor_0')
	browser.find_element_by_xpath('.//body[last()]').send_keys(answer)
	browser.switch_to.default_content() #get back main iframe
	rest(2,3)
#	browser.find_element_by_name('md').click()#匿名回答
	__authcode()	#check authcode
	browser.find_element_by_xpath('//a[@alog-action="qb-ans-sb"]').click()
	rest(2,5)
	count = count + 1
	res = check_result()
	if 0==res:
		status_count = status_count - 1
		print('回答成功')
		coutinue_count = 0
		failed_status = 0
		return 0
	else:
		print('回答失败')
		failed_status = 1
		return -1

	
	

#返回值 1表示出现异常，0表示回答成功，-1表示回答失败
def answer_map(question_url):
	global coutinue_count,failed_status,status_count,count
	title = get_title(question_url)
	#过滤失效的问题
	if not title:
		print('链接打不开')
		return 1
	print('回答地图:%s'%(title))
	
	browser.get(question_url)
	rest(3,5)
	#如果回答过了，立即返回
	try:
		browser.find_element_by_xpath('//a[@alog-action="qb-modify-ans"]')
		print('回答过了')
		return -1
	except Exception:
		pass
			
	try:
#		time.sleep(1)
		browser.find_element_by_xpath('//a[@alog-action="qb-ans-sb"]') 
	except Exception:
		try:
			browser.find_element_by_id("answer-bar").click()
#			time.sleep(1)
		except Exception:
			print('无法回答')
#		time.sleep(1)
			return -1
	rest(1,3)
	browser.find_element_by_id("edui14").click()
	ele_iframe = browser.find_element_by_class_name("ui-dialog-content-iframe")
	browser.switch_to_frame(ele_iframe)
	rest(1,3)
	browser.find_element_by_xpath('//input[@value="插入地图"]').click()
		
	rest(1,3)
	#check authcode
	__authcode()
	#匿名
	browser.switch_to_default_content()
	browser.find_element_by_name('md').click()
	rest(1,3)
	browser.find_element_by_xpath('//a[@alog-action="qb-ans-sb"]').click()
	rest(3,5)
	count = count + 1
	res = check_result()
	if 0==res:
#		answers_quest.append(question_url)
		print('回答成功')
		hide_my_anwser()
		rest(3,5)
		res = finish_answer(question_url)
		failed_status = 0
		return 0
	else:
		print('回答失败')
		failed_status = 1
	coutinue_count = coutinue_count + 1




###########################################end


#########################################################
#alog-action="qb-modify-ans"   完善
#alog-action="qb-ans-sb"	提交修改
def finish_answer(question_url):
	global failed_status,coutinue_count,count
	
	title = get_title(question_url)
	#过滤失效的问题
	if not title:
		print('链接打不开')
		return 1
	browser.get(question_url)
	rest(1,3)
	try:
		#完善回答
		browser.find_element_by_xpath('//a[@alog-action="qb-modify-ans"]').click() 
	except NoSuchElementException:
		print('无法完善回答')
		return 1
	answer = ""
	answer = genate_anwser()
	print('anwser:')
	print(answer)
	browser.switch_to_frame('ueditor_0')
	rest(1,3)
	ele = browser.find_element_by_xpath('.//body[last()]')#回答框
	ele.clear()
	trest(1,3)
	ele.send_keys(answer)
	rest(1,3)


	browser.switch_to.default_content()
	__authcode()
	try:
		browser.find_element_by_xpath('//a[@alog-action="qb-ans-sb"]').click()
	finally:
		pass
#	input()
	res = check_result()
	count = count + 1
	if 0==res:
		print('完善回答成功')
		answers_quest.append(question_url)
		rest(1,3)
		return 0
	else:
		print('完善回答失败')
		return -1
		#从干扰文案列表删除失效文案
		l = len(oth_list)
		oth_text = oth_list[(ans_count-1)%l]
		oth_list.remove(oth_text)
	coutinue_count = coutinue_count + 1
###########################################




def do_other_thing(url):
	choice = random.randint(0,3)
	if 0==choice:
		print(u'抽奖')
		lottery()
	elif 1==choice:
		praise(url)
	elif 2==choice:
		check_message()
	else:
		print(u'休息一下')
		rest(choice,10)
		
def praise(url):
	browser.get(url)
	index = random.randint(0,4)
	try:
		rest(1,3)
		#target = browser.find_element_by_xpath('i[@class="iknow-qb_home_icons"]')
		target = browser.find_element_by_css_selector(".iknow-qb_home_icons")
		browser.execute_script("arguments[0].scrollIntoView();", target) 
		#rest(3,5)
		tag_list = browser.find_elements_by_xpath('//a[@class="related-link"]')
		tag_list[random.randint(0,4)].click()
		rest(1,2)
		browser.switch_to_window(browser.window_handles[1])
		rest(3,5)
		if random.randint(0,2)%2:
			print(u'点赞')
			browser.find_element_by_xpath('//span[@alog-action="qb-zan-btn"]').click()
		else:
			print(u'评论')
			browser.find_element_by_xpath('//span[@alog-action="qb-comment-btn"]').click()
			pre = ''.join(random.sample(string.ascii_letters + string.digits, 8))
			str = pre+' '+comment_text[random.randint(0,4)]
			
			js = 'document.querySelectorAll("textarea")[0].style.display="block";'
			browser.execute_script(js)
			item = browser.find_element_by_xpath('//div[@class="comment-area no-comment"]//textarea')
#			print(item.is_displayed())
			rest(1,3)
			item.send_keys(str)
			browser.find_element_by_xpath('//a[@alog-action="qb-comment-submit"]').click()
	except Exception as e:
		print(e)
	finally:
		rest(3,5)
		print(browser.window_handles)
		for handler in browser.window_handles:
			if browser.window_handles.index(handler)>=1:
				browser.switch_to_window(handler)
				browser.close()
		browser.switch_to_window(browser.window_handles[0])
		print(browser.current_window_handle)
		
def lottery():
	browser.get(home)
	
	try:
#		browser.find_element_by_xpath('//a[@class="menu-right-list-item user-center"]').click()
#		browser.switch_to_window(browser.window_handles[1])
		rest(1,3)
		browser.find_element_by_xpath('//a[@class="btn-36-green-new grid"]').click()
		rest(1,3)
		browser.switch_to_window(browser.window_handles[1])
		browser.find_element_by_xpath('//a[@href="/shop/lottery"]').click()
		rest(1,3)
		browser.find_element_by_xpath('//a[@class="control-btn one-try-btn"]').click()
		rest(5,10)
		browser.find_element_by_xpath('//a[@class="lucky-btn lucky-btn-ok"]').click()
	except Exception as e:
		print(u'抽奖失败')
	finally:
		rest(1,3)
		print(browser.window_handles)
		for handler in browser.window_handles:
			if browser.window_handles.index(handler)>=1:
				browser.switch_to_window(handler)
				browser.close()
		browser.switch_to_window(browser.window_handles[0])
#		print(browser.current_window_handle)


def check_message():
	print('check message')
	rest(3,5)
	browser.find_element_by_xpath('//a[@href="/ihome/notice/all"]').click()
	browser.switch_to_window(browser.window_handles[1])
	btn1 = browser.find_element_by_xpath('//a[@href="/ihome/notice/reply"]')
	btn2 = browser.find_element_by_xpath('//a[@href="/ihome/notice/sysmsg"]')
	btn3 = browser.find_element_by_xpath('//a[@href="/ihome/notice/ask"]')
	index = random.randint(1,3)
	btn = None
	if index==1:
		print(u'查看回答消息')
		btn = btn1
	elif index==2:
		print(u'查看系统消息')
		btn = btn2
	elif index==3:
		print(u'查看提问消息')
		btn = btn3
	else:
		print(u'休息一下吧')
		rest(3,8)
	if btn:
		btn.click()
		rest(1,3)
	target = browser.find_element_by_xpath('//div[@class="footer-title"]')
	scroll_bar(target)
	for handler in browser.window_handles:
		if browser.window_handles.index(handler)>=1:
			browser.switch_to_window(handler)
			browser.close()
	browser.switch_to_window(browser.window_handles[0])
	
	
	
	
def scroll_bar(target):
	browser.execute_script("arguments[0].scrollIntoView();", target) #拖动到可见的元素去
		
def login():
	if BDUSS:
		log_by_cookie(BDUSS)
	elif username:
		__login(username,password)
	else:
		print('error')
		exit(-1)

	
#######################################程序开始	
comment_text = ['are you sure?','anything else?',"may be you are right",' some plase may be better',"I don't think so"]
#zes_0001@163.com一zhu1234
#username = 'zes_0001@163.com'
#password = 'zhu1234'
#username = "依恋晨qz"
#password = "{[(zhu1993)]}"	

#username = 'zhu_jhsreee@163.com'
#password = 'zhu1234'
#username = 'lua8688@outlook.com'
#password = '{[(zhu1993)]}'
username = read_para('username')
password = read_para('password')
BDUSS = read_para('BDUSS')
suffix = '?fr=qlquick&entry=qb_list_default'
#八级号
#BDUSS='3FPbE80YVZzVDVtRXZmVVdaSkJnTTJYOWhpdnNNcGk1QlRUbUFYV0EyQkZOTzlZSVFBQUFBJCQAAAAAAAAAAAEAAABqsu8j0sDBtbO~cXoAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAEWnx1hFp8dYV'

#八级号2
#BDUSS='RSZUtXTTRqSkFMQmRZVy1uckpKZXZ6dXNWbjc4T2praHc0TUFMOVNlNG1jfkJZSVFBQUFBJCQAAAAAAAAAAAEAAAB5guCp38~fz9~P1ajBywAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACbmyFgm5shYNj'

#普通5级BDUSS = 'VHVVBuTkJtS3Q5eWFUbkE4MlB0NjJDLUQwc3RNekxjd3ZOdkZvd2JEQ3ltLTlZSVFBQUFBJCQAAAAAAAAAAAEAAADdzL5lAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAALIOyFiyDshYMX'
question_list = []
answers_quest = []
real_list = []
valid_wenan = []
home = 'https://zhidao.baidu.com/ihome/answer?tab=1'
baidu = 'https://zhidao.baidu.com'	
browser = webdriver.Chrome()



login()
url = 'https://zhidao.baidu.com/question/79600056.html'

for i in range(0,5):
	do_other_thing(url)
input()
exit(0)

#0表示成功，1表示失败
failed_status = 0
coutinue_count = 0
authcode = 0


#####################################
#添加养号问题
add_question(question_list)

#获取需要回答地图的链接
read_file(real_list,'real.txt')
print('需要回答地图%d'%(len(real_list)))
status_count = 0



'''
####################################只养号



while True:
	if len(question_list):
		quest = question_list.pop()
		quest = quest + suffix
		if check_question(quest):
			continue
		rand = random.randint(2,5)
		time.sleep(rand)
	
		if anwser_question(quest):
			continue
	else:
		print('正在抓取问题')
		add_question(question_list)
		
	if not count%30:
		print('按下回车键继续')
		input()
		print('开始回答')
save(answers_quest)
print('正在保存问题')
time.sleep(3)

#######################################

'''




#######################################养号+答地图
count = 0
while True:	
	rand = random.randint(3,10)
#	time.sleep(rand)
	
	if authcode==1 or failed_status==1 or (failed_status==0 and rand%2) or coutinue_count>2 :
		authcode = 0
		if len(question_list):
			quest = question_list.pop() + suffix
			if anwser_question(quest):
				rest(1,3)
				coutinue_count = 0
				continue
					
		else:
			print('正在抓取问题')
			add_question(question_list)
	else:
		if rand==5:
			do_other_thing()
		if not len(real_list):
			break
		quest = real_list.pop() + suffix
		rest(3,5)
		res = answer_map(quest)
		if res==-1:
			rest(3,5)
		else:
			rest(1,3)
		continue
	
	
	if not (count+1)%20:
		print('休息一下吧，回答了20条了.按下回车键继续')
		print('保存问题请输入：s')
		res = input()
		if res=='s':
			save(answers_quest)
		print('回答继续')
	

save(answers_quest)
print('正在保存问题')
rest(3,5)
	

#######################################

	
